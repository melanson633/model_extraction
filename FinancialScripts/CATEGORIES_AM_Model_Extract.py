import os
import datetime
from openpyxl import load_workbook
import csv

# Path to the Excel file (now a directory for batch processing)
excel_path = r"C:\Users\mmelanson\.cursor-tutor\projects\AM Model Extraction\Data\_MODEL"
# Path to directory where CSV should be saved (now the TEST output folder)
output_dir = r"C:\Users\mmelanson\.cursor-tutor\projects\AM Model Extraction\Output\FINANCIALS"

# Path to the mapping dictionary (still in project root)
mapping_csv = os.path.join(os.path.dirname(__file__), 'unique_categories_dictionary.csv')
# Path to log unmapped categories (still in project root)
unmapped_log_path = os.path.join(os.path.dirname(__file__), 'unmapped_categories.log')

# Load mapping dictionary
category_mapping = {}
with open(mapping_csv, 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        key = row['MODEL_UniqueCategory'].strip()
        val = row['EXTRACTION_MAPPING'].strip()
        category_mapping[key] = val

def get_first_of_month(date_val):
    """
    Given a date (datetime.date or datetime.datetime), return the first day of that month.
    """
    if isinstance(date_val, (datetime.date, datetime.datetime)):
        return datetime.date(date_val.year, date_val.month, 1)
    # Try to parse string date
    try:
        dt = datetime.datetime.strptime(str(date_val), "%m/%d/%y")
        return datetime.date(dt.year, dt.month, 1)
    except Exception:
        return date_val  # fallback, should not happen

def process_excel_file(file_path, output_dir, category_mapping, unmapped_categories):
    original_filename = os.path.splitext(os.path.basename(file_path))[0]
    # Extract the text immediately preceding the first '-' character
    if '-' in original_filename:
        prefix = original_filename.split('-')[0].strip()
    else:
        prefix = original_filename.strip()
    csv_filename = f"model_csv_{prefix}.csv"
    csv_path = os.path.join(output_dir, csv_filename)

    # Derive Source column value:
    filename_before_dot = original_filename.split('.', 1)[0]
    source_value = filename_before_dot.split(' - ')[0].strip()

    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb["Model"] if "Model" in wb.sheetnames else wb[wb.sheetnames[0]]

    # 1. Find the cell containing "Categories"
    categories_cell = None
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column):
        for cell in row:
            if cell.value == "Categories":
                categories_cell = cell
                break
        if categories_cell:
            break

    if categories_cell is None:
        print(f"Could not find a cell containing 'Categories' in {file_path}.")
        return

    categories_row = categories_cell.row
    categories_col = categories_cell.column  # integer index

    # 2. Identify the first date column: immediately to the right of Categories
    first_date_col = categories_col + 1

    # 3. Determine date columns
    date_columns = []
    col_idx = first_date_col
    while True:
        val = ws.cell(row=categories_row, column=col_idx).value
        if val is None:
            break
        parsed_date = None
        if isinstance(val, (datetime.date, datetime.datetime)):
            parsed_date = val.date() if isinstance(val, datetime.datetime) else val
        else:
            try:
                parsed_date = datetime.datetime.strptime(str(val), "%m/%d/%y").date()
            except Exception:
                break
        if parsed_date:
            # Always use first of month for output
            first_of_month = get_first_of_month(parsed_date)
            date_columns.append((col_idx, first_of_month))
            col_idx += 1
        else:
            break

    if not date_columns:
        print(f"No date columns found in {file_path}.")
        return

    # 4. Find "Change in Cash"
    change_in_cash_row = None
    for r in range(categories_row+1, ws.max_row+1):
        val = ws.cell(row=r, column=categories_col).value
        if val == "Change in Cash":
            change_in_cash_row = r
            break

    if change_in_cash_row is None:
        print(f"Could not find 'Change in Cash' row in {file_path}.")
        return

    # Categories rows
    start_category_row = categories_row + 1
    end_category_row = change_in_cash_row - 1

    # 5. Extract data
    data = []
    for r in range(start_category_row, end_category_row + 1):
        category_name = ws.cell(row=r, column=categories_col).value
        if not category_name:
            continue
        category_str = str(category_name).strip()
        # Remap category if mapping exists
        mapped_category = category_mapping.get(category_str, category_str)
        if category_str not in category_mapping:
            unmapped_categories.add((os.path.basename(file_path), category_str))
        for (dc, period) in date_columns:
            amount_val = ws.cell(row=r, column=dc).value
            # Convert to float, round to 2 decimals, and ensure output is a number
            try:
                amount_val = float(amount_val)
                amount_val = round(amount_val, 2)
            except (ValueError, TypeError):
                amount_val = 0.0
            data.append((mapped_category, period, amount_val))

    # 6. Write to CSV with an additional "Source" column
    # Filter out rows where amount is 0.0
    filtered_data = [row for row in data if row[2] != 0.0]
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["CATEGORY", "MONTH", "AMOUNT", "PROPERTY"])
        for row_data in filtered_data:
            category, period, amount = row_data
            writer.writerow([category, period.strftime("%Y-%m-%d"), amount, source_value])

    print(f"Data extracted and saved to {csv_path}")

# Batch process all .xlsx files in the directory
unmapped_categories = set()
for fname in os.listdir(excel_path):
    if not fname.lower().endswith('.xlsx'):
        continue
    fpath = os.path.join(excel_path, fname)
    process_excel_file(fpath, output_dir, category_mapping, unmapped_categories)

# Write unmapped categories to log file
if unmapped_categories:
    with open(unmapped_log_path, 'w', encoding='utf-8') as logf:
        for file_name, category in sorted(unmapped_categories):
            logf.write(f"{file_name}: {category}\n")
    print(f"Unmapped categories logged to {unmapped_log_path}")
