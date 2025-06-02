import os
import csv
import datetime
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# Directory containing Excel files
EXCEL_DIR = r"C:\Users\mmelanson\.cursor-tutor\projects\AM Model Extraction\Data\_MODEL"

# Generate a timestamped filename for the output CSV
now = datetime.datetime.now()
timestamp = now.strftime("%Y%m%d_%H%M%S")
output_filename = f'unique_categories_dictionary_{timestamp}.csv'
OUTPUT_CSV = os.path.join(os.path.dirname(__file__), output_filename)


def extract_categories_from_file(filepath):
    """
    Extracts category names from the 'Categories' column up to 'Change in Cash' in the given Excel file.
    Returns a set of category names. Handles errors gracefully.
    """
    categories = set()
    try:
        wb = load_workbook(filename=filepath, data_only=True, read_only=True)
        # Try all sheets, but prefer 'Model' if present
        sheetnames = wb.sheetnames
        if 'Model' in sheetnames:
            ws = wb['Model']
        else:
            ws = wb[sheetnames[0]]

        # 1. Find the cell containing 'Categories'
        categories_cell = None
        for row in ws.iter_rows(min_row=1, max_col=ws.max_column):
            for cell in row:
                if str(cell.value).strip().lower() == 'categories':
                    categories_cell = cell
                    break
            if categories_cell:
                break
        if not categories_cell:
            return categories  # No anchor, skip file
        categories_row = categories_cell.row
        categories_col = categories_cell.column

        # 2. Find the row with 'Change in Cash' in the same column
        change_in_cash_row = None
        for r in range(categories_row + 1, ws.max_row + 1):
            val = ws.cell(row=r, column=categories_col).value
            if val is not None and str(val).strip().lower() == 'change in cash':
                change_in_cash_row = r
                break
        if not change_in_cash_row:
            return categories  # No end anchor, skip file

        # 3. Extract all non-empty, non-numeric, non-header values between anchors
        for r in range(categories_row + 1, change_in_cash_row):
            val = ws.cell(row=r, column=categories_col).value
            if val is None:
                continue
            val_str = str(val).strip()
            if not val_str:
                continue
            # Skip if value is numeric or looks like a date
            try:
                float(val_str.replace(',', ''))
                continue  # skip numbers
            except ValueError:
                pass
            # Skip if value is the header or end anchor
            if val_str.lower() in ['categories', 'change in cash']:
                continue
            categories.add(val_str)
    except (InvalidFileException, PermissionError, OSError, KeyError):
        # Skip unreadable, password-protected, or corrupted files
        pass
    except Exception as e:
        # Log or print if needed, but continue
        pass
    return categories


def main():
    """
    Batch process all .xlsx files in the specified directory, extract unique categories, and export to CSV.
    """
    all_categories = set()
    print(f"Starting to process files in: {EXCEL_DIR}")
    files_to_process = [fname for fname in os.listdir(EXCEL_DIR) if fname.lower().endswith('.xlsx')]
    if not files_to_process:
        print("No .xlsx files found in the specified directory.")
        return
    
    print(f"Found {len(files_to_process)} .xlsx files to process.")

    for fname in files_to_process:
        fpath = os.path.join(EXCEL_DIR, fname)
        print(f"Processing file: {fpath}...")
        try:
            cats = extract_categories_from_file(fpath)
            all_categories.update(cats)
            print(f"Successfully processed and extracted categories from: {fname}")
        except KeyboardInterrupt:
            print(f"Processing of {fname} was interrupted by the user (KeyboardInterrupt).")
            # Re-raise the interrupt to stop the script if desired, or handle differently
            raise
        except Exception as e:
            # This will catch errors from extract_categories_from_file if they aren't handled internally
            # or if an error occurs before/after the internal try-except in that function.
            print(f"An error occurred while processing file {fpath}: {e}")
            # Decide if you want to continue to the next file or stop.
            # For now, we'll continue, similar to the original script's silent error handling.
            pass # Continue to the next file

    if not all_categories:
        print("No categories were extracted from any files.")
    else:
        # Sort for readability
        unique_categories = sorted(all_categories)
        # Write to CSV
        try:
            with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['UniqueCategory'])
                for cat in unique_categories:
                    writer.writerow([cat])
            print(f"Extracted {len(unique_categories)} unique categories. Output saved to {OUTPUT_CSV}")
        except IOError as e:
            print(f"Error writing to CSV file {OUTPUT_CSV}: {e}")
        except Exception as e:
            print(f"An unexpected error occurred during CSV writing: {e}")


if __name__ == '__main__':
    main() 